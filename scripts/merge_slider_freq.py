"""Clean rebuild: create new workbook with merged data."""
import re
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from copy import copy

EXCEL_PATH = "data/359098880_1777624158BAfsef.xlsx"

OCR_DATA = {
    10: (102.2, 3.19, [2, 6, 11, 9, 4]),
    11: (135.6, 4.24, [1, 2, 3, 6, 20]),
    12: (114.2, 3.57, [2, 5, 8, 6, 11]),
    13: (122, 3.81, [1, 3, 7, 10, 11]),
    14: (115.5, 3.61, [3, 2, 7, 11, 9]),
    15: (124, 3.88, [2, 2, 4, 12, 12]),
    21: (108.7, 3.4, [4, 3, 6, 15, 4]),
    22: (59.3, 1.85, [22, 4, 1, 4, 1]),
    23: (87.3, 2.73, [7, 7, 8, 8, 2]),
    24: (111.7, 3.49, [1, 5, 10, 11, 5]),
    25: (108.4, 3.39, [2, 4, 10, 10, 6]),
    26: (129.4, 4.04, [3, 2, 1, 10, 16]),
    27: (144, 4.5, [0, 0, 1, 11, 20]),
}

BINS = ["1–1.8分", "1.9–2.6分", "2.7–3.4分", "3.5–4.2分", "4.3–5分"]


def extract_q_num(cell_text: str) -> int | None:
    if "第" in cell_text and "题" in cell_text:
        m = re.search(r"第(\d+)题", cell_text)
        if m:
            return int(m.group(1))
    return None


def main():
    # Read original data
    wb_old = openpyxl.load_workbook(EXCEL_PATH)
    ws_old = wb_old["sheet1"]
    all_rows: list[list] = []
    for row in ws_old.iter_rows(min_row=1, max_row=ws_old.max_row, min_col=1, max_col=ws_old.max_column):
        all_rows.append([cell.value for cell in row])
    wb_old.close()

    # Identify original question boundaries
    # For slider questions, find: title row, summary row, then skip to next question
    clean_rows: list[list] = []
    i = 0
    while i < len(all_rows):
        row_data = all_rows[i]
        cell_a = str(row_data[0] or "")
        q_num = extract_q_num(cell_a)

        if q_num is not None and "滑动条" in cell_a and q_num in OCR_DATA:
            # Slider question with OCR data
            # Keep title row
            clean_rows.append(list(row_data))
            i += 1

            # Keep summary row (总分值/平均值)
            if i < len(all_rows) and ("总分值" in str(all_rows[i][0] or "")):
                clean_rows.append(list(all_rows[i]))
                i += 1

            # Skip all rows until next question title or blank area
            while i < len(all_rows):
                next_cell = str(all_rows[i][0] or "")
                next_q = extract_q_num(next_cell)
                if next_q is not None:
                    break  # next question found
                # Also stop at first truly blank row after some content
                if all(v is None or str(v).strip() == "" for v in all_rows[i]):
                    i += 1
                    break
                i += 1

            # Add correct frequency data
            total, avg, freqs = OCR_DATA[q_num]
            total_resp = sum(freqs)
            clean_rows.append(["分段", "频数", "比例"])
            for bin_label, freq in zip(BINS, freqs):
                pct = f"{freq / total_resp * 100:.2f}%" if total_resp else "0%"
                clean_rows.append([bin_label, freq, pct])
            clean_rows.append(["本题有效填写人次", total_resp, ""])
            clean_rows.append([None, None, None])
            continue

        # Not a target slider question — skip any pre-existing frequency data
        # that was added by previous runs (detect "分段" header or bin labels near slider)
        clean_rows.append(list(row_data))
        i += 1

    # Write to brand new workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "sheet1"

    for row_idx, row_data in enumerate(clean_rows, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            if val is not None:
                ws_new.cell(row=row_idx, column=col_idx, value=val)

    wb_new.save(EXCEL_PATH)
    print(f"Done. {len(clean_rows)} rows written to new workbook.")


if __name__ == "__main__":
    main()
